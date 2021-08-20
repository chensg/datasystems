import time
import openpyxl
from pathlib import Path
from pathlib import PurePath
from os import walk

from models import DimDate
from models import FactMaintenanceContractorPayment
from dbhelper import DatabaseHelper
import log

# set up logger
logger = log.setup_logger('root')

class MainApp(object):

    def __init__(self):
        self.data = []
        self.sheet = None

    def extract(self, xlsx_file):
        # Step 1 Extract: use openpyxl library to open xls file and extract data from the file
        logger.debug(f"Starting Step 1 Extracting data from files")
        # xlsx_file = Path('data', '41091_maintenanceStaffLogbookV1a.xlsx')
        wb_obj = openpyxl.load_workbook(xlsx_file)
        self.sheet = wb_obj.active
        logger.debug(f"we find {self.sheet.max_row} rows and {self.sheet.max_column} columns in file {xlsx_file}")
        logger.debug(f"Step 1 finished")

    def transform(self):
        # Step 2 Transform: the transform detail is located in the function transform of Model class
        # and we will put transformed model into a list

        logger.debug(f"Starting Step 2 transforming data")
        self.data = []
        for i, row in enumerate(self.sheet.iter_rows(values_only = True)):
            if i == 0:
                # we ignore the first row, which is the head titles
                continue

            if row[0] is None:
                # may be the last row
                continue
            
            logger.debug(f"transform data: {row}")
            payment_record = FactMaintenanceContractorPayment()
            payment_record.transform(row)
            self.data.append(payment_record)

        logger.debug(f"Step 2 finished")

    def load(self):
        # Step 3 Load: we will use model function save to load the data into database

        logger.debug(f"Starting Step 3 loading data into database")
        for model in self.data:
            logger.debug(f"load data into database: {model}")
            model.save()
        logger.debug(f"Step 3 finished")

    def mainLoop(self):

        logger.debug("Check if there are any new files in the incoming directory")

        # first of all, we need to clearn fact table for the demo purpose
        DatabaseHelper.getInstance().cleanFactTable()
        incoming_path = Path('data', 'incoming')
        done_path = Path('data', 'done')
        filenames = next(walk(incoming_path), (None, None, []))[2]
        logger.debug(f"find {len(filenames)} new files in incoming directory")
        for filename in filenames:
            xlsx_file = Path(incoming_path, filename)
            # step 1
            self.extract(xlsx_file)
            # step 2
            self.transform()
            # step 3
            self.load()
            Path(xlsx_file).rename(Path(done_path, filename))

        # close database connection
        DatabaseHelper.getInstance().close()

def main():
    # create an instance of MainApp
    main = MainApp()
    while True:
        main.mainLoop()
        logger.debug("sleeping for 60 seconds")
        time.sleep(60)

if __name__ == "__main__":
    main()
